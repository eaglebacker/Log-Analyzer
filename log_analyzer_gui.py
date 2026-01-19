"""
Log Analyzer GUI Application
Reads log files and exports to Excel with filtered error lines.
Supports multiple custom filter tabs.
"""

import json
import os
import re
import sys
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from datetime import datetime, date

import csv

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.hyperlink import Hyperlink
except ImportError:
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.hyperlink import Hyperlink

# Try to import tkinterdnd2 for drag and drop support
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    DND_AVAILABLE = True
except ImportError:
    DND_AVAILABLE = False

# Excel row limit (max rows per worksheet)
EXCEL_MAX_ROWS = 1048576


class LogAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Log Analyzer")
        self.root.geometry("850x700")
        self.root.resizable(True, True)
        self.root.minsize(750, 600)

        # Save config on window close
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        # Variables
        self.input_file = tk.StringVar()
        self.output_file = tk.StringVar()

        # Filters: dict of {filter_name: [keywords]}
        self.filters = {}
        self.selected_filter = None

        # Default filters
        self.default_filters = {
            "Errors": ["Error", "Exception", "Fatal", "Critical", "Fail", "Failed"],
            "Client Command": ["Client Command"],
            "AET warning": ["AET warning"],
            "Bigfoot General": [
                "Current Software Version",
                "CNozzleTipDefinitionModel Pressure",
                "Set Drop Drive Frequency from drop delay records",
                "Setting Defanning",
                "Set Sample Offset",
                "Set Boost Time",
                "Set Boost Overpressure",
                "Set Area Scalar",
                "Set Droplet Camera Position",
                "Set Sort Straight Deflection",
                "Set Wash Time",
                "Set Wash Pressure",
                "Set Wash Backflush Time",
                "Setting Drop Delay:"
            ]
        }

        # Tab colors for Excel (cycling through these)
        self.tab_colors = [
            ("C00000", "Red"),
            ("4472C4", "Blue"),
            ("70AD47", "Green"),
            ("ED7D31", "Orange"),
            ("7030A0", "Purple"),
            ("00B0F0", "Cyan"),
            ("FFC000", "Yellow"),
        ]

        # Concatenate mode
        self.concat_mode_var = tk.BooleanVar(value=False)
        self.selected_files = []

        # Date color palette for Excel
        self.date_colors = [
            "E6F3FF",  # Light Blue
            "FFE6F0",  # Light Pink
            "FFF4E6",  # Light Orange
            "F0FFE6",  # Light Green
            "F3E6FF",  # Light Purple
            "FFFFE6",  # Light Yellow
            "E6FFFF",  # Light Cyan
            "FFE6E6",  # Light Red
        ]

        # Load configuration before creating widgets
        config_loaded = self.load_config()

        self.create_widgets()

        # Only load defaults if no config was loaded
        if not config_loaded:
            self.load_default_filters()
        else:
            self.populate_filter_list()

    def get_config_path(self):
        """
        Get the path to the config file.
        Handles both script and PyInstaller executable scenarios.
        """
        if getattr(sys, 'frozen', False):
            # Running as compiled executable
            app_dir = Path(sys.executable).parent
        else:
            # Running as script
            app_dir = Path(__file__).parent

        return app_dir / "config.json"

    def validate_geometry(self, geometry_string):
        """
        Validate a Tkinter geometry string format.
        Expected format: WIDTHxHEIGHT or WIDTHxHEIGHT+X+Y
        Returns True if valid, False otherwise.
        """
        # Pattern: WIDTHxHEIGHT or WIDTHxHEIGHT+X+Y or WIDTHxHEIGHT-X-Y (mixed signs)
        pattern = r'^\d+x\d+([+-]\d+[+-]\d+)?$'
        return bool(re.match(pattern, geometry_string))

    def load_config(self):
        """
        Load configuration from JSON file.
        Falls back to defaults if file doesn't exist or is corrupted.
        Returns True if config was loaded, False if using defaults.
        """
        config_path = self.get_config_path()

        # If config doesn't exist, use defaults
        if not config_path.exists():
            return False

        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)

            # Validate version (for future compatibility)
            version = config.get('version', '1.0')

            # Load filters (with validation)
            if 'filters' in config and isinstance(config['filters'], dict):
                # Validate filter structure
                valid_filters = {}
                for name, keywords in config['filters'].items():
                    if isinstance(keywords, list) and all(isinstance(k, str) for k in keywords):
                        valid_filters[name] = keywords

                if valid_filters:
                    self.filters = valid_filters

            # Load file paths (with validation)
            if 'paths' in config:
                paths = config['paths']
                if 'last_input_file' in paths:
                    input_path = paths['last_input_file']
                    # Only restore if file still exists
                    if input_path and os.path.isfile(input_path):
                        self.input_file.set(input_path)

                if 'last_output_file' in paths:
                    output_path = paths['last_output_file']
                    if output_path:
                        self.output_file.set(output_path)

            # Load concat mode
            if 'concat_mode' in config:
                self.concat_mode_var.set(config['concat_mode'])

            # Load window geometry (with validation)
            if 'window' in config and 'geometry' in config['window']:
                geometry = config['window']['geometry']
                # Validate geometry string format (WIDTHxHEIGHT+X+Y or WIDTHxHEIGHT)
                if self.validate_geometry(geometry):
                    self.root.geometry(geometry)

            return True

        except (json.JSONDecodeError, KeyError, TypeError, ValueError) as e:
            # Config file is corrupted - use defaults
            print(f"Warning: Could not load config file: {e}")
            return False

    def save_config(self):
        """
        Save current application state to JSON config file.
        Called automatically after any state change.
        Silently fails if unable to save (non-critical operation).
        """
        try:
            config_path = self.get_config_path()

            # Get current window geometry
            geometry = self.root.geometry()

            # Build config structure
            config = {
                'version': '1.0',
                'concat_mode': self.concat_mode_var.get(),
                'filters': dict(self.filters),  # Create a copy
                'paths': {
                    'last_input_file': self.input_file.get(),
                    'last_output_file': self.output_file.get()
                },
                'window': {
                    'geometry': geometry
                }
            }

            # Write atomically (write to temp file, then rename)
            temp_path = config_path.with_suffix('.json.tmp')
            with open(temp_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2, ensure_ascii=False)

            # Atomic rename (prevents corruption if write is interrupted)
            temp_path.replace(config_path)

        except (IOError, OSError, TypeError) as e:
            # Silently fail - config saving is non-critical
            # Could optionally log this for debugging
            pass

    def create_widgets(self):
        # Main container with padding
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky="nsew")

        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)

        # === Input File Section ===
        dnd_text = " (or drag & drop)" if DND_AVAILABLE else ""
        input_frame = ttk.LabelFrame(main_frame, text=f"1. Select Log File{dnd_text}", padding="10")
        input_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        input_frame.columnconfigure(0, weight=1)

        input_entry = ttk.Entry(input_frame, textvariable=self.input_file, width=70)
        input_entry.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        self.input_entry = input_entry  # Store reference for drag and drop

        browse_input_btn = ttk.Button(input_frame, text="Browse...", command=self.browse_input)
        browse_input_btn.grid(row=0, column=1)

        # Concatenate mode checkbox
        concat_frame = ttk.Frame(input_frame)
        concat_frame.grid(row=1, column=0, columnspan=2, sticky="w", pady=(5, 0))

        self.concat_checkbox = ttk.Checkbutton(
            concat_frame,
            text="Enable Concatenate Mode (select up to 10 files)",
            variable=self.concat_mode_var,
            command=self.on_concat_mode_toggle
        )
        self.concat_checkbox.pack(side="left")

        self.selected_files_label = ttk.Label(concat_frame, text="", foreground="blue")
        self.selected_files_label.pack(side="left", padx=(10, 0))

        # Register drag and drop on the entire input frame section
        if DND_AVAILABLE:
            # Register drop target on the frame and all child widgets
            input_frame.drop_target_register(DND_FILES)
            input_frame.dnd_bind('<<Drop>>', self.on_file_drop)
            input_entry.drop_target_register(DND_FILES)
            input_entry.dnd_bind('<<Drop>>', self.on_file_drop)
            concat_frame.drop_target_register(DND_FILES)
            concat_frame.dnd_bind('<<Drop>>', self.on_file_drop)

        # === Output File Section ===
        output_frame = ttk.LabelFrame(main_frame, text="2. Save Output As", padding="10")
        output_frame.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        output_frame.columnconfigure(0, weight=1)

        output_entry = ttk.Entry(output_frame, textvariable=self.output_file, width=70)
        output_entry.grid(row=0, column=0, sticky="ew", padx=(0, 10))

        browse_output_btn = ttk.Button(output_frame, text="Browse...", command=self.browse_output)
        browse_output_btn.grid(row=0, column=1)

        # === Filter Tabs Section ===
        filters_frame = ttk.LabelFrame(main_frame, text="3. Filter Tabs (each creates a separate Excel sheet)", padding="10")
        filters_frame.grid(row=2, column=0, sticky="nsew", pady=(0, 10))
        filters_frame.columnconfigure(0, weight=1)
        filters_frame.columnconfigure(1, weight=2)
        filters_frame.rowconfigure(1, weight=1)
        main_frame.rowconfigure(2, weight=1)

        # --- Left side: Filter list ---
        filter_list_frame = ttk.Frame(filters_frame)
        filter_list_frame.grid(row=0, column=0, rowspan=2, sticky="nsew", padx=(0, 10))
        filter_list_frame.rowconfigure(1, weight=1)

        ttk.Label(filter_list_frame, text="Filter Tabs:", font=("Segoe UI", 9, "bold")).grid(row=0, column=0, sticky="w")

        # Filter listbox
        filter_listbox_frame = ttk.Frame(filter_list_frame)
        filter_listbox_frame.grid(row=1, column=0, sticky="nsew", pady=(5, 5))
        filter_listbox_frame.rowconfigure(0, weight=1)

        self.filter_listbox = tk.Listbox(filter_listbox_frame, height=10, width=25, exportselection=False)
        self.filter_listbox.grid(row=0, column=0, sticky="nsew")
        self.filter_listbox.bind("<<ListboxSelect>>", self.on_filter_select)

        filter_scroll = ttk.Scrollbar(filter_listbox_frame, orient="vertical", command=self.filter_listbox.yview)
        filter_scroll.grid(row=0, column=1, sticky="ns")
        self.filter_listbox.configure(yscrollcommand=filter_scroll.set)

        # Filter buttons
        filter_btn_frame = ttk.Frame(filter_list_frame)
        filter_btn_frame.grid(row=2, column=0, sticky="ew", pady=(5, 0))

        ttk.Button(filter_btn_frame, text="+ Add Tab", command=self.add_filter, width=12).grid(row=0, column=0, padx=(0, 5))
        ttk.Button(filter_btn_frame, text="- Remove", command=self.remove_filter, width=12).grid(row=0, column=1)

        # --- Right side: Keywords for selected filter ---
        keywords_frame = ttk.Frame(filters_frame)
        keywords_frame.grid(row=0, column=1, rowspan=2, sticky="nsew")
        keywords_frame.columnconfigure(0, weight=1)
        keywords_frame.rowconfigure(2, weight=1)

        self.keywords_label = ttk.Label(keywords_frame, text="Keywords for selected tab:", font=("Segoe UI", 9, "bold"))
        self.keywords_label.grid(row=0, column=0, sticky="w")

        # Add keyword row
        add_kw_frame = ttk.Frame(keywords_frame)
        add_kw_frame.grid(row=1, column=0, sticky="ew", pady=(5, 5))
        add_kw_frame.columnconfigure(0, weight=1)

        self.new_keyword_entry = ttk.Entry(add_kw_frame, width=30)
        self.new_keyword_entry.grid(row=0, column=0, sticky="ew", padx=(0, 5))
        self.new_keyword_entry.bind("<Return>", lambda e: self.add_keyword())

        ttk.Button(add_kw_frame, text="Add", command=self.add_keyword, width=8).grid(row=0, column=1, padx=(0, 5))
        ttk.Button(add_kw_frame, text="Remove", command=self.remove_keyword, width=8).grid(row=0, column=2)

        # Keywords listbox
        kw_listbox_frame = ttk.Frame(keywords_frame)
        kw_listbox_frame.grid(row=2, column=0, sticky="nsew")
        kw_listbox_frame.columnconfigure(0, weight=1)
        kw_listbox_frame.rowconfigure(0, weight=1)

        self.keywords_listbox = tk.Listbox(kw_listbox_frame, height=8, selectmode=tk.EXTENDED)
        self.keywords_listbox.grid(row=0, column=0, sticky="nsew")

        kw_scroll = ttk.Scrollbar(kw_listbox_frame, orient="vertical", command=self.keywords_listbox.yview)
        kw_scroll.grid(row=0, column=1, sticky="ns")
        self.keywords_listbox.configure(yscrollcommand=kw_scroll.set)

        # Reset button
        ttk.Button(keywords_frame, text="Reset All to Defaults", command=self.load_default_filters).grid(
            row=3, column=0, sticky="w", pady=(10, 0)
        )

        # === Analyze Button ===
        analyze_btn = ttk.Button(main_frame, text="Analyze Log File", command=self.analyze, style="Accent.TButton")
        analyze_btn.grid(row=3, column=0, pady=10, ipady=10, sticky="ew")

        style = ttk.Style()
        style.configure("Accent.TButton", font=("Segoe UI", 11, "bold"))

        # === Status Section ===
        status_frame = ttk.LabelFrame(main_frame, text="Status", padding="10")
        status_frame.grid(row=4, column=0, sticky="ew")
        status_frame.columnconfigure(0, weight=1)

        self.status_label = ttk.Label(status_frame, text="Ready. Select a log file to begin.", foreground="gray")
        self.status_label.grid(row=0, column=0, sticky="w")

        self.progress = ttk.Progressbar(status_frame, mode="indeterminate")
        self.progress.grid(row=1, column=0, sticky="ew", pady=(10, 0))

    def browse_input(self):
        if self.concat_mode_var.get():
            # Multi-file selection mode
            filenames = filedialog.askopenfilenames(
                title="Select Log Files (up to 10)",
                filetypes=[
                    ("Log files", "*.log"),
                    ("Text files", "*.txt"),
                    ("All files", "*.*")
                ]
            )
            if filenames:
                if len(filenames) > 10:
                    messagebox.showerror("Too Many Files",
                        "Please select up to 10 files maximum.")
                    return

                self.selected_files = list(filenames)
                self.update_file_selection_display()

                # Auto-generate output filename based on first file
                if self.selected_files:
                    input_path = Path(self.selected_files[0])
                    output_path = input_path.parent / f"concatenated_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    self.output_file.set(str(output_path))
                self.save_config()
        else:
            # Single-file selection mode (existing logic)
            filename = filedialog.askopenfilename(
                title="Select Log File",
                filetypes=[
                    ("Log files", "*.log"),
                    ("Text files", "*.txt"),
                    ("All files", "*.*")
                ]
            )
            if filename:
                self.input_file.set(filename)
                input_path = Path(filename)
                output_path = input_path.parent / f"{input_path.stem}_analysis.xlsx"
                self.output_file.set(str(output_path))
                self.save_config()

    def browse_output(self):
        filename = filedialog.asksaveasfilename(
            title="Save Output As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if filename:
            self.output_file.set(filename)
            self.save_config()

    def on_file_drop(self, event):
        """Handle files dropped onto the input field."""
        # Parse the dropped data - tkinterdnd2 returns file paths
        # Files with spaces are enclosed in curly braces: {C:/path/with spaces/file.log}
        dropped_data = event.data

        # Parse file paths from the dropped data
        files = []
        # Handle curly brace notation for paths with spaces
        if '{' in dropped_data:
            # Extract paths from curly braces
            import re
            brace_paths = re.findall(r'\{([^}]+)\}', dropped_data)
            files.extend(brace_paths)
            # Also get paths not in braces
            remaining = re.sub(r'\{[^}]+\}', '', dropped_data).strip()
            if remaining:
                files.extend(remaining.split())
        else:
            files = dropped_data.split()

        # Filter to only existing files (not directories)
        valid_files = [f for f in files if os.path.isfile(f)]

        if not valid_files:
            messagebox.showwarning("Invalid Drop", "No valid files were dropped.")
            return

        if self.concat_mode_var.get():
            # Multi-file mode
            if len(valid_files) > 10:
                messagebox.showerror("Too Many Files",
                    "Please drop up to 10 files maximum.")
                return

            self.selected_files = valid_files
            self.update_file_selection_display()

            # Auto-generate output filename
            input_path = Path(self.selected_files[0])
            output_path = input_path.parent / f"concatenated_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            self.output_file.set(str(output_path))
        else:
            # Single-file mode - use first file only
            if len(valid_files) > 1:
                messagebox.showinfo("Single File Mode",
                    f"Single file mode is active. Using first file:\n{os.path.basename(valid_files[0])}\n\n"
                    "Enable 'Concatenate Mode' to use multiple files.")

            self.input_file.set(valid_files[0])
            input_path = Path(valid_files[0])
            output_path = input_path.parent / f"{input_path.stem}_analysis.xlsx"
            self.output_file.set(str(output_path))

        self.save_config()

    def on_concat_mode_toggle(self):
        """Handle concatenate mode toggle."""
        self.selected_files = []
        self.input_file.set("")
        self.update_file_selection_display()

    def update_file_selection_display(self):
        """Update UI to show selected files count."""
        if self.concat_mode_var.get() and self.selected_files:
            count = len(self.selected_files)
            self.selected_files_label.configure(text=f"{count} file(s) selected")
            # Show summary in input field
            if count == 1:
                self.input_file.set(self.selected_files[0])
            else:
                self.input_file.set(f"{count} files selected")
        else:
            self.selected_files_label.configure(text="")

    # --- Filter Management ---
    def add_filter(self):
        """Add a new filter tab."""
        dialog = FilterNameDialog(self.root, "New Filter Tab", existing_names=list(self.filters.keys()))
        if dialog.result:
            filter_name = dialog.result
            self.filters[filter_name] = []
            self.filter_listbox.insert(tk.END, filter_name)
            # Select the new filter
            self.filter_listbox.selection_clear(0, tk.END)
            self.filter_listbox.selection_set(tk.END)
            self.on_filter_select(None)
            self.save_config()

    def remove_filter(self):
        """Remove the selected filter tab."""
        selection = self.filter_listbox.curselection()
        if not selection:
            messagebox.showinfo("No Selection", "Please select a filter tab to remove.")
            return

        filter_name = self.filter_listbox.get(selection[0])
        if messagebox.askyesno("Confirm", f"Remove filter tab '{filter_name}'?"):
            del self.filters[filter_name]
            self.filter_listbox.delete(selection[0])
            self.keywords_listbox.delete(0, tk.END)
            self.selected_filter = None
            self.keywords_label.configure(text="Keywords for selected tab:")
            self.save_config()

    def on_filter_select(self, event):
        """Handle filter selection change."""
        selection = self.filter_listbox.curselection()
        if not selection:
            return

        filter_name = self.filter_listbox.get(selection[0])
        self.selected_filter = filter_name
        self.keywords_label.configure(text=f"Keywords for '{filter_name}':")

        # Update keywords listbox
        self.keywords_listbox.delete(0, tk.END)
        for keyword in self.filters.get(filter_name, []):
            self.keywords_listbox.insert(tk.END, keyword)

    # --- Keyword Management ---
    def add_keyword(self):
        """Add a keyword to the selected filter."""
        if not self.selected_filter:
            messagebox.showwarning("No Filter Selected", "Please select a filter tab first.")
            return

        keyword = self.new_keyword_entry.get().strip()
        if not keyword:
            messagebox.showwarning("Empty", "Please enter a keyword.")
            return

        if keyword in self.filters[self.selected_filter]:
            messagebox.showwarning("Duplicate", f"'{keyword}' is already in this filter.")
            return

        self.filters[self.selected_filter].append(keyword)
        self.keywords_listbox.insert(tk.END, keyword)
        self.new_keyword_entry.delete(0, tk.END)
        self.save_config()

    def remove_keyword(self):
        """Remove selected keywords from the current filter."""
        if not self.selected_filter:
            messagebox.showwarning("No Filter Selected", "Please select a filter tab first.")
            return

        selection = self.keywords_listbox.curselection()
        if not selection:
            messagebox.showinfo("No Selection", "Please select keyword(s) to remove.")
            return

        for idx in reversed(selection):
            keyword = self.keywords_listbox.get(idx)
            self.filters[self.selected_filter].remove(keyword)
            self.keywords_listbox.delete(idx)
        self.save_config()

    def load_default_filters(self):
        """Reset to default filters."""
        self.filters = {name: list(keywords) for name, keywords in self.default_filters.items()}
        self.filter_listbox.delete(0, tk.END)
        for filter_name in self.filters:
            self.filter_listbox.insert(tk.END, filter_name)

        # Select first filter
        if self.filters:
            self.filter_listbox.selection_set(0)
            self.on_filter_select(None)
        self.save_config()

    def populate_filter_list(self):
        """
        Populate the filter listbox with loaded filters.
        Used when loading from config instead of defaults.
        """
        self.filter_listbox.delete(0, tk.END)
        for filter_name in self.filters:
            self.filter_listbox.insert(tk.END, filter_name)

        # Select first filter if any exist
        if self.filters:
            self.filter_listbox.selection_set(0)
            self.on_filter_select(None)

    def on_closing(self):
        """
        Handle window close event.
        Save config one final time before closing.
        """
        self.save_config()
        self.root.destroy()

    # --- Analysis ---
    def update_status(self, message, color="black"):
        self.status_label.configure(text=message, foreground=color)
        self.root.update_idletasks()

    def analyze(self):
        if self.concat_mode_var.get():
            # Multi-file mode validation
            if not self.selected_files:
                messagebox.showerror("Error", "Please select log files to concatenate.")
                return

            # Validate all files exist
            missing_files = [f for f in self.selected_files if not os.path.isfile(f)]
            if missing_files:
                messagebox.showerror("Error",
                    f"File(s) not found:\n" + "\n".join([os.path.basename(f) for f in missing_files]))
                return

            input_path = None  # Not used in concat mode
        else:
            # Single-file mode validation (existing logic)
            input_path = self.input_file.get().strip()

            if not input_path:
                messagebox.showerror("Error", "Please select an input log file.")
                return

            if not os.path.isfile(input_path):
                messagebox.showerror("Error", f"Input file not found:\n{input_path}")
                return

        output_path = self.output_file.get().strip()

        if not output_path:
            messagebox.showerror("Error", "Please specify an output file location.")
            return

        # Check that at least one filter has keywords
        filters_with_keywords = {name: kws for name, kws in self.filters.items() if kws}
        if not filters_with_keywords:
            messagebox.showerror("Error", "Please add at least one filter with keywords.")
            return

        self.progress.start()
        thread = threading.Thread(
            target=self.run_analysis,
            args=(input_path, output_path, filters_with_keywords.copy())
        )
        thread.start()

    def run_analysis(self, input_path, output_path, filters):
        try:
            if self.concat_mode_var.get():
                # Multi-file concatenation mode
                self.update_status(f"Reading {len(self.selected_files)} log files...", "blue")

                # Read and tag all files
                tagged_lines = self.read_and_tag_log_files(self.selected_files)

                if not tagged_lines:
                    self.progress.stop()
                    messagebox.showerror("Error", "No valid lines found in selected files.")
                    return

                # Sort chronologically
                self.update_status("Sorting lines chronologically...", "blue")
                sorted_lines = self.sort_and_merge_lines(tagged_lines)
                total_lines = len(sorted_lines)

                # Warn if no timestamps
                if all(t[0] is None for t in sorted_lines):
                    messagebox.showwarning("No Timestamps",
                        "Could not parse any timestamps from the selected files. "
                        "Lines will be displayed in file order without chronological sorting.")

                # Apply filters to sorted lines
                self.update_status(f"Read {total_lines:,} lines. Applying filters...", "blue")
                filter_results = {}
                for filter_name, keywords in filters.items():
                    matches = self.filter_tagged_lines(sorted_lines, keywords)
                    filter_results[filter_name] = {
                        "keywords": keywords,
                        "matches": matches
                    }

                # Check if row limit exceeded
                max_data_rows = EXCEL_MAX_ROWS - 1  # Account for header row
                export_mode = "excel"  # Default to normal Excel export
                created_files = None

                if total_lines > max_data_rows:
                    # Show dialog on main thread
                    self.progress.stop()
                    dialog = RowLimitDialog(self.root, total_lines)
                    export_mode = dialog.result

                    if export_mode is None:
                        # User cancelled
                        self.update_status("Export cancelled by user.", "gray")
                        return

                    self.progress.start()

                # Create output based on user choice
                if export_mode == "csv":
                    self.update_status(f"Creating CSV files...", "blue")
                    created_files = self.create_csv_report_concat(sorted_lines, filter_results, output_path, self.selected_files)
                elif export_mode == "split":
                    self.update_status(f"Creating Excel file with split sheets...", "blue")
                    self.create_excel_report_concat_split(sorted_lines, filter_results, output_path, self.selected_files)
                else:
                    self.update_status(f"Creating Excel file with {len(filters)} filter tab(s)...", "blue")
                    self.create_excel_report_concat(sorted_lines, filter_results, output_path, self.selected_files)

                self.progress.stop()

                # Build summary message
                summary_parts = []
                for name, data in filter_results.items():
                    summary_parts.append(f"{name}: {len(data['matches']):,}")

                self.update_status(
                    f"Done! {' | '.join(summary_parts)} (Total lines: {total_lines:,})",
                    "green"
                )

                if export_mode == "csv" and created_files:
                    file_list = "\n".join([f"  - {f.name}" for f in created_files])
                    if messagebox.askyesno("Success", f"CSV export complete!\n\nCreated files:\n{file_list}\n\nOpen the folder now?"):
                        os.startfile(created_files[0].parent)
                else:
                    if messagebox.askyesno("Success", f"Analysis complete!\n\nOpen the Excel file now?"):
                        os.startfile(output_path)
            else:
                # Single-file mode (existing logic)
                self.update_status("Reading log file...", "blue")
                all_lines = self.read_log_file(input_path)
                total_lines = len(all_lines)

                self.update_status(f"Read {total_lines:,} lines. Applying filters...", "blue")

                # Apply each filter
                filter_results = {}
                for filter_name, keywords in filters.items():
                    matches = self.filter_lines(all_lines, keywords)
                    filter_results[filter_name] = {
                        "keywords": keywords,
                        "matches": matches
                    }

                # Check if row limit exceeded
                max_data_rows = EXCEL_MAX_ROWS - 1  # Account for header row
                export_mode = "excel"  # Default to normal Excel export
                created_files = None

                if total_lines > max_data_rows:
                    # Show dialog on main thread
                    self.progress.stop()
                    dialog = RowLimitDialog(self.root, total_lines)
                    export_mode = dialog.result

                    if export_mode is None:
                        # User cancelled
                        self.update_status("Export cancelled by user.", "gray")
                        return

                    self.progress.start()

                # Create output based on user choice
                if export_mode == "csv":
                    self.update_status(f"Creating CSV files...", "blue")
                    created_files = self.create_csv_report(all_lines, filter_results, output_path, input_path)
                elif export_mode == "split":
                    self.update_status(f"Creating Excel file with split sheets...", "blue")
                    self.create_excel_report_split(all_lines, filter_results, output_path, input_path)
                else:
                    self.update_status(f"Creating Excel file with {len(filters)} filter tab(s)...", "blue")
                    self.create_excel_report(all_lines, filter_results, output_path, input_path)

                self.progress.stop()

                # Build summary message
                summary_parts = []
                for name, data in filter_results.items():
                    summary_parts.append(f"{name}: {len(data['matches']):,}")

                self.update_status(
                    f"Done! {' | '.join(summary_parts)} (Total lines: {total_lines:,})",
                    "green"
                )

                if export_mode == "csv" and created_files:
                    file_list = "\n".join([f"  - {f.name}" for f in created_files])
                    if messagebox.askyesno("Success", f"CSV export complete!\n\nCreated files:\n{file_list}\n\nOpen the folder now?"):
                        os.startfile(created_files[0].parent)
                else:
                    if messagebox.askyesno("Success", f"Analysis complete!\n\nOpen the Excel file now?"):
                        os.startfile(output_path)

        except Exception as e:
            self.progress.stop()
            self.update_status(f"Error: {str(e)}", "red")
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")

    def read_log_file(self, file_path):
        encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    lines = f.readlines()
                return [line.rstrip('\n\r') for line in lines]
            except UnicodeDecodeError:
                continue
        raise ValueError("Could not read file with supported encodings.")

    def parse_timestamp(self, line):
        """
        Extract timestamp from log line.
        Returns datetime object or None if parsing fails.
        """
        # Pattern: YYYY-MM-DD HH:MM:SS.FFFF at start of line
        pattern = r'^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{4})'
        match = re.match(pattern, line)

        if match:
            timestamp_str = match.group(1)
            try:
                return datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S.%f')
            except ValueError:
                return None
        return None

    def read_and_tag_log_files(self, file_paths):
        """
        Read multiple log files and return tagged lines.
        Returns list of tuples: (timestamp, filename, line_content, original_line_num)
        """
        tagged_lines = []

        for file_path in file_paths:
            try:
                lines = self.read_log_file(file_path)
                filename = os.path.basename(file_path)

                for line_num, line in enumerate(lines, start=1):
                    timestamp = self.parse_timestamp(line)
                    tagged_lines.append((timestamp, filename, line, line_num))
            except Exception as e:
                self.update_status(f"Warning: Could not read {os.path.basename(file_path)}: {e}", "orange")
                continue

        return tagged_lines

    def sort_and_merge_lines(self, tagged_lines):
        """
        Sort tagged lines by timestamp chronologically.
        Lines with unparseable timestamps go to the end.
        """
        # Separate lines with valid timestamps from those without
        with_timestamp = [t for t in tagged_lines if t[0] is not None]
        without_timestamp = [t for t in tagged_lines if t[0] is None]

        # Sort lines with timestamps
        with_timestamp.sort(key=lambda x: x[0])

        # Concatenate: sorted lines first, then lines without timestamps
        return with_timestamp + without_timestamp

    def filter_lines(self, lines, keywords):
        """Filter lines that contain any of the keywords."""
        matches = []
        keywords_lower = [kw.lower() for kw in keywords]
        for line_num, line in enumerate(lines, start=1):
            line_lower = line.lower()
            if any(keyword in line_lower for keyword in keywords_lower):
                matches.append((line_num, line))
        return matches

    def filter_tagged_lines(self, tagged_lines, keywords):
        """
        Filter tagged lines that contain any of the keywords.
        Returns list of tagged tuples that match.
        """
        matches = []
        keywords_lower = [kw.lower() for kw in keywords]

        for tagged_line in tagged_lines:
            timestamp, filename, line, line_num = tagged_line
            line_lower = line.lower()
            if any(keyword in line_lower for keyword in keywords_lower):
                matches.append(tagged_line)

        return matches

    def extract_unique_dates(self, tagged_lines):
        """Extract sorted list of unique dates from tagged lines."""
        dates = set()
        for timestamp, filename, line, line_num in tagged_lines:
            if timestamp:
                dates.add(timestamp.date())
        return sorted(dates)

    def build_date_color_map(self, unique_dates):
        """Create mapping of date to background color."""
        date_color_map = {}
        for idx, date_obj in enumerate(unique_dates):
            color = self.date_colors[idx % len(self.date_colors)]
            date_color_map[date_obj] = color
        return date_color_map

    def create_excel_report(self, all_lines, filter_results, output_path, source_file):
        wb = Workbook()

        header_font = Font(bold=True, color="FFFFFF")

        # --- Sheet 1: All Logs ---
        ws_all = wb.active
        ws_all.title = "All Logs"

        all_logs_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
        ws_all['A1'] = "Line #"
        ws_all['B1'] = "Log Content"
        ws_all['A1'].font = header_font
        ws_all['A1'].fill = all_logs_fill
        ws_all['B1'].font = header_font
        ws_all['B1'].fill = all_logs_fill

        for idx, line in enumerate(all_lines, start=2):
            ws_all[f'A{idx}'] = idx - 1
            ws_all[f'B{idx}'] = line

        ws_all.column_dimensions['A'].width = 10
        ws_all.column_dimensions['B'].width = 150

        # --- Filter Sheets ---
        color_idx = 0
        link_font = Font(color="0563C1", underline="single")  # Blue underlined hyperlink style

        for filter_name, data in filter_results.items():
            ws = wb.create_sheet(title=filter_name[:31])  # Excel sheet names max 31 chars

            # Get color for this tab
            color_hex = self.tab_colors[color_idx % len(self.tab_colors)][0]
            header_fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            color_idx += 1

            ws['A1'] = "Original Line #"
            ws['B1'] = "Log Content"
            ws['A1'].font = header_font
            ws['A1'].fill = header_fill
            ws['B1'].font = header_font
            ws['B1'].fill = header_fill

            for idx, (line_num, line) in enumerate(data['matches'], start=2):
                cell = ws[f'A{idx}']
                cell.value = line_num
                # Add hyperlink to corresponding row in All Logs (line_num + 1 for header)
                all_logs_row = line_num + 1
                cell.hyperlink = f"#'All Logs'!A{all_logs_row}"
                cell.font = link_font
                ws[f'B{idx}'] = line

            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 150

        # --- Summary Sheet ---
        ws_summary = wb.create_sheet(title="Summary")
        summary_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

        ws_summary['A1'] = "Metric"
        ws_summary['B1'] = "Value"
        ws_summary['A1'].font = header_font
        ws_summary['A1'].fill = summary_fill
        ws_summary['B1'].font = header_font
        ws_summary['B1'].fill = summary_fill

        row = 2
        ws_summary[f'A{row}'] = "Source File"
        ws_summary[f'B{row}'] = source_file
        row += 1

        ws_summary[f'A{row}'] = "Total Lines"
        ws_summary[f'B{row}'] = len(all_lines)
        row += 2

        ws_summary[f'A{row}'] = "Filter Results"
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1

        for filter_name, data in filter_results.items():
            ws_summary[f'A{row}'] = f"  {filter_name}"
            ws_summary[f'B{row}'] = f"{len(data['matches']):,} lines ({len(data['matches']) / len(all_lines) * 100:.2f}%)"
            row += 1

            ws_summary[f'A{row}'] = f"    Keywords"
            ws_summary[f'B{row}'] = ", ".join(data['keywords'])
            row += 1

        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 80

        wb.save(output_path)

    def create_excel_report_concat(self, sorted_tagged_lines, filter_results, output_path, source_files):
        """
        Create Excel report for concatenated multi-file analysis.
        Optimized for large files with bulk operations.
        """
        self.update_status("Creating Excel workbook...", "blue")
        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF")

        # Build date color mapping
        unique_dates = self.extract_unique_dates(sorted_tagged_lines)
        date_color_map = self.build_date_color_map(unique_dates)

        # Pre-create fill objects (reuse instead of creating new ones each time)
        date_fills = {date_obj: PatternFill(start_color=color, end_color=color, fill_type="solid")
                      for date_obj, color in date_color_map.items()}
        separator_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        separator_font = Font(bold=True, size=11)

        # --- Sheet 1: All Logs (Optimized) ---
        self.update_status("Writing All Logs sheet...", "blue")
        ws_all = wb.active
        ws_all.title = "All Logs"

        # Headers
        all_logs_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
        ws_all.append(["Line #", "Date", "Source File", "Log Content"])
        for col in ['A1', 'B1', 'C1', 'D1']:
            ws_all[col].font = header_font
            ws_all[col].fill = all_logs_fill

        # Build all rows first, track positions for formatting
        rows_to_add = []
        color_positions = {}  # {row_num: date_obj}
        separator_positions = []  # [(row_num, date_str)]
        all_logs_row_map = {}  # {id(tagged_line): final_row} for hyperlinks
        current_row = 2
        current_date = None
        progress_counter = 0
        progress_interval = 5000

        for idx, tagged_line in enumerate(sorted_tagged_lines, start=1):
            timestamp, filename, line, original_line_num = tagged_line
            line_date = timestamp.date() if timestamp else None

            # Track separator positions
            if line_date and line_date != current_date:
                if current_date is not None:
                    separator_positions.append((current_row, str(line_date)))
                    current_row += 1
                current_date = line_date

            # Build row data
            rows_to_add.append([idx, str(line_date) if line_date else "N/A", filename, line])

            # Track color position (only for Date column B)
            if line_date:
                color_positions[current_row] = line_date

            # Map this tagged line to its All Logs row (for hyperlinks in filter sheets)
            all_logs_row_map[id(tagged_line)] = current_row

            current_row += 1

            # Progress update
            progress_counter += 1
            if progress_counter % progress_interval == 0:
                self.update_status(f"Processing All Logs: {progress_counter:,}/{len(sorted_tagged_lines):,} lines...", "blue")

        # Bulk append all rows (much faster)
        self.update_status("Writing All Logs rows to Excel...", "blue")
        for row_data in rows_to_add:
            ws_all.append(row_data)

        # Apply separators (after rows are added)
        self.update_status("Adding date separators to All Logs...", "blue")
        for sep_row, date_str in separator_positions:
            ws_all.insert_rows(sep_row)
            ws_all[f'A{sep_row}'] = f"=== Date: {date_str} ==="
            ws_all.merge_cells(f'A{sep_row}:D{sep_row}')
            ws_all[f'A{sep_row}'].font = separator_font
            ws_all[f'A{sep_row}'].fill = separator_fill

        # Apply color coding (only to Date column B for performance)
        self.update_status("Applying date colors to All Logs...", "blue")
        for row_num, date_obj in color_positions.items():
            if date_obj in date_fills:
                # Adjust row numbers after separator insertions
                adjusted_row = row_num + len([s for s in separator_positions if s[0] <= row_num])
                ws_all[f'B{adjusted_row}'].fill = date_fills[date_obj]

        # Set column widths
        ws_all.column_dimensions['A'].width = 10
        ws_all.column_dimensions['B'].width = 12
        ws_all.column_dimensions['C'].width = 30
        ws_all.column_dimensions['D'].width = 120

        # --- Filter Sheets (Optimized) ---
        color_idx = 0
        link_font = Font(color="0563C1", underline="single")  # Blue underlined hyperlink style

        for filter_name, data in filter_results.items():
            self.update_status(f"Writing filter sheet: {filter_name}...", "blue")
            ws = wb.create_sheet(title=filter_name[:31])

            # Get color for this filter tab
            tab_color_hex = self.tab_colors[color_idx % len(self.tab_colors)][0]
            header_fill = PatternFill(start_color=tab_color_hex, end_color=tab_color_hex, fill_type="solid")
            color_idx += 1

            # Headers
            ws.append(["Line #", "Date", "Source File", "Log Content"])
            for col in ['A1', 'B1', 'C1', 'D1']:
                ws[col].font = header_font
                ws[col].fill = header_fill

            # Build rows and track positions, including hyperlink info
            rows_to_add = []
            hyperlink_info = []  # [(filter_row, all_logs_row), ...]
            color_positions = {}
            separator_positions = []
            current_row = 2
            current_date = None

            for idx, tagged_line in enumerate(data['matches'], start=1):
                timestamp, filename, line, original_line_num = tagged_line
                line_date = timestamp.date() if timestamp else None

                # Track separator positions
                if line_date and line_date != current_date:
                    if current_date is not None:
                        separator_positions.append((current_row, str(line_date)))
                        current_row += 1
                    current_date = line_date

                # Build row data
                rows_to_add.append([idx, str(line_date) if line_date else "N/A", filename, line])

                # Track hyperlink info - map filter sheet row to All Logs row
                all_logs_row = all_logs_row_map.get(id(tagged_line))
                if all_logs_row:
                    hyperlink_info.append((current_row, all_logs_row))

                # Track color position
                if line_date:
                    color_positions[current_row] = line_date

                current_row += 1

            # Bulk append rows
            for row_data in rows_to_add:
                ws.append(row_data)

            # Apply separators
            for sep_row, date_str in separator_positions:
                ws.insert_rows(sep_row)
                ws[f'A{sep_row}'] = f"=== Date: {date_str} ==="
                ws.merge_cells(f'A{sep_row}:D{sep_row}')
                ws[f'A{sep_row}'].font = separator_font
                ws[f'A{sep_row}'].fill = separator_fill

            # Apply hyperlinks to Line # column
            for filter_row, all_logs_row in hyperlink_info:
                # Adjust for separators inserted in this filter sheet
                adjusted_filter_row = filter_row + len([s for s in separator_positions if s[0] <= filter_row])
                cell = ws[f'A{adjusted_filter_row}']
                cell.hyperlink = f"#'All Logs'!A{all_logs_row}"
                cell.font = link_font

            # Apply color coding (only Date column)
            for row_num, date_obj in color_positions.items():
                if date_obj in date_fills:
                    adjusted_row = row_num + len([s for s in separator_positions if s[0] <= row_num])
                    ws[f'B{adjusted_row}'].fill = date_fills[date_obj]

            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 120

        # --- Summary Sheet ---
        self.update_status("Creating summary sheet...", "blue")
        ws_summary = wb.create_sheet(title="Summary")
        summary_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

        ws_summary['A1'] = "Metric"
        ws_summary['B1'] = "Value"
        ws_summary['A1'].font = header_font
        ws_summary['A1'].fill = summary_fill
        ws_summary['B1'].font = header_font
        ws_summary['B1'].fill = summary_fill

        row = 2
        ws_summary[f'A{row}'] = "Mode"
        ws_summary[f'B{row}'] = "Multi-File Concatenation"
        row += 1

        ws_summary[f'A{row}'] = "Number of Files"
        ws_summary[f'B{row}'] = len(source_files)
        row += 2

        ws_summary[f'A{row}'] = "Source Files"
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1

        for file_path in source_files:
            ws_summary[f'A{row}'] = f"  {os.path.basename(file_path)}"
            row += 1

        row += 1
        ws_summary[f'A{row}'] = "Total Lines"
        ws_summary[f'B{row}'] = len(sorted_tagged_lines)
        row += 2

        ws_summary[f'A{row}'] = "Date Range"
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1

        if unique_dates:
            ws_summary[f'A{row}'] = "  Start Date"
            ws_summary[f'B{row}'] = str(unique_dates[0])
            row += 1
            ws_summary[f'A{row}'] = "  End Date"
            ws_summary[f'B{row}'] = str(unique_dates[-1])
            row += 2

        ws_summary[f'A{row}'] = "Filter Results"
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1

        for filter_name, data in filter_results.items():
            ws_summary[f'A{row}'] = f"  {filter_name}"
            ws_summary[f'B{row}'] = f"{len(data['matches']):,} lines"
            row += 1

            ws_summary[f'A{row}'] = f"    Keywords"
            ws_summary[f'B{row}'] = ", ".join(data['keywords'])
            row += 1

        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 80

        self.update_status("Saving Excel file...", "blue")
        wb.save(output_path)

    def create_csv_report(self, all_lines, filter_results, output_path, source_file):
        """
        Create CSV report for single-file analysis.
        Creates multiple CSV files: one for all logs and one for each filter.
        """
        base_path = Path(output_path)
        base_name = base_path.stem
        output_dir = base_path.parent

        created_files = []

        # --- All Logs CSV ---
        all_logs_path = output_dir / f"{base_name}_all_logs.csv"
        self.update_status("Writing All Logs CSV...", "blue")

        with open(all_logs_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["Line #", "Log Content"])
            for idx, line in enumerate(all_lines, start=1):
                writer.writerow([idx, line])

        created_files.append(all_logs_path)

        # --- Filter CSVs ---
        for filter_name, data in filter_results.items():
            # Sanitize filter name for filename
            safe_name = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in filter_name)
            filter_path = output_dir / f"{base_name}_{safe_name}.csv"

            self.update_status(f"Writing {filter_name} CSV...", "blue")

            with open(filter_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["Original Line #", "Log Content"])
                for line_num, line in data['matches']:
                    writer.writerow([line_num, line])

            created_files.append(filter_path)

        # --- Summary CSV ---
        summary_path = output_dir / f"{base_name}_summary.csv"
        with open(summary_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["Metric", "Value"])
            writer.writerow(["Source File", source_file])
            writer.writerow(["Total Lines", len(all_lines)])
            writer.writerow([])
            writer.writerow(["Filter Results", ""])
            for filter_name, data in filter_results.items():
                pct = len(data['matches']) / len(all_lines) * 100 if all_lines else 0
                writer.writerow([filter_name, f"{len(data['matches']):,} lines ({pct:.2f}%)"])
                writer.writerow(["  Keywords", ", ".join(data['keywords'])])

        created_files.append(summary_path)

        return created_files

    def create_csv_report_concat(self, sorted_tagged_lines, filter_results, output_path, source_files):
        """
        Create CSV report for concatenated multi-file analysis.
        Creates multiple CSV files: one for all logs and one for each filter.
        """
        base_path = Path(output_path)
        base_name = base_path.stem
        output_dir = base_path.parent

        created_files = []

        # --- All Logs CSV ---
        all_logs_path = output_dir / f"{base_name}_all_logs.csv"
        self.update_status("Writing All Logs CSV...", "blue")

        with open(all_logs_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["Line #", "Date", "Source File", "Log Content"])
            for idx, (timestamp, filename, line, original_line_num) in enumerate(sorted_tagged_lines, start=1):
                line_date = str(timestamp.date()) if timestamp else "N/A"
                writer.writerow([idx, line_date, filename, line])

        created_files.append(all_logs_path)

        # --- Filter CSVs ---
        for filter_name, data in filter_results.items():
            # Sanitize filter name for filename
            safe_name = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in filter_name)
            filter_path = output_dir / f"{base_name}_{safe_name}.csv"

            self.update_status(f"Writing {filter_name} CSV...", "blue")

            with open(filter_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["Line #", "Date", "Source File", "Log Content"])
                for idx, (timestamp, filename, line, original_line_num) in enumerate(data['matches'], start=1):
                    line_date = str(timestamp.date()) if timestamp else "N/A"
                    writer.writerow([idx, line_date, filename, line])

            created_files.append(filter_path)

        # --- Summary CSV ---
        summary_path = output_dir / f"{base_name}_summary.csv"
        unique_dates = self.extract_unique_dates(sorted_tagged_lines)

        with open(summary_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["Metric", "Value"])
            writer.writerow(["Mode", "Multi-File Concatenation"])
            writer.writerow(["Number of Files", len(source_files)])
            writer.writerow([])
            writer.writerow(["Source Files", ""])
            for file_path in source_files:
                writer.writerow(["", os.path.basename(file_path)])
            writer.writerow([])
            writer.writerow(["Total Lines", len(sorted_tagged_lines)])
            if unique_dates:
                writer.writerow(["Start Date", str(unique_dates[0])])
                writer.writerow(["End Date", str(unique_dates[-1])])
            writer.writerow([])
            writer.writerow(["Filter Results", ""])
            for filter_name, data in filter_results.items():
                writer.writerow([filter_name, f"{len(data['matches']):,} lines"])
                writer.writerow(["  Keywords", ", ".join(data['keywords'])])

        created_files.append(summary_path)

        return created_files

    def create_excel_report_split(self, all_lines, filter_results, output_path, source_file):
        """
        Create Excel report with All Logs split across multiple sheets.
        Used when total lines exceed Excel's row limit.
        """
        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF")

        # Calculate how many sheets needed for All Logs
        max_data_rows = EXCEL_MAX_ROWS - 1  # -1 for header row
        total_lines = len(all_lines)
        num_sheets = (total_lines + max_data_rows - 1) // max_data_rows  # Ceiling division

        self.update_status(f"Splitting All Logs across {num_sheets} sheets...", "blue")

        all_logs_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")

        # --- All Logs Sheets (split) ---
        for sheet_num in range(num_sheets):
            start_idx = sheet_num * max_data_rows
            end_idx = min((sheet_num + 1) * max_data_rows, total_lines)

            if sheet_num == 0:
                ws = wb.active
                ws.title = "All Logs"
            else:
                ws = wb.create_sheet(title=f"All Logs ({sheet_num + 1})")

            # Headers
            ws['A1'] = "Line #"
            ws['B1'] = "Log Content"
            ws['A1'].font = header_font
            ws['A1'].fill = all_logs_fill
            ws['B1'].font = header_font
            ws['B1'].fill = all_logs_fill

            # Add data rows
            for row_offset, line_idx in enumerate(range(start_idx, end_idx)):
                excel_row = row_offset + 2  # +2 for 1-indexed and header
                ws[f'A{excel_row}'] = line_idx + 1  # Original line number (1-indexed)
                ws[f'B{excel_row}'] = all_lines[line_idx]

            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 150

            self.update_status(f"Wrote All Logs sheet {sheet_num + 1}/{num_sheets}...", "blue")

        # --- Filter Sheets ---
        color_idx = 0
        link_font = Font(color="0563C1", underline="single")  # Blue underlined hyperlink style

        # Helper function to get sheet name and row for a line number
        def get_all_logs_location(line_num):
            """Return (sheet_name, row) for a given line number in split All Logs."""
            sheet_idx = (line_num - 1) // max_data_rows
            row_in_sheet = (line_num - 1) % max_data_rows + 2  # +2 for header and 1-indexing
            if sheet_idx == 0:
                return "All Logs", row_in_sheet
            else:
                return f"All Logs ({sheet_idx + 1})", row_in_sheet

        for filter_name, data in filter_results.items():
            matches = data['matches']
            filter_total = len(matches)

            # Check if this filter also needs splitting
            if filter_total > max_data_rows:
                filter_num_sheets = (filter_total + max_data_rows - 1) // max_data_rows
                for sheet_num in range(filter_num_sheets):
                    start_idx = sheet_num * max_data_rows
                    end_idx = min((sheet_num + 1) * max_data_rows, filter_total)

                    if sheet_num == 0:
                        sheet_title = filter_name[:31]
                    else:
                        base_title = filter_name[:26]  # Leave room for " (N)"
                        sheet_title = f"{base_title} ({sheet_num + 1})"

                    ws = wb.create_sheet(title=sheet_title)

                    color_hex = self.tab_colors[color_idx % len(self.tab_colors)][0]
                    header_fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

                    ws['A1'] = "Original Line #"
                    ws['B1'] = "Log Content"
                    ws['A1'].font = header_font
                    ws['A1'].fill = header_fill
                    ws['B1'].font = header_font
                    ws['B1'].fill = header_fill

                    for row_offset, match_idx in enumerate(range(start_idx, end_idx)):
                        excel_row = row_offset + 2
                        line_num, line = matches[match_idx]
                        cell = ws[f'A{excel_row}']
                        cell.value = line_num
                        # Add hyperlink to All Logs
                        target_sheet, target_row = get_all_logs_location(line_num)
                        cell.hyperlink = f"#'{target_sheet}'!A{target_row}"
                        cell.font = link_font
                        ws[f'B{excel_row}'] = line

                    ws.column_dimensions['A'].width = 15
                    ws.column_dimensions['B'].width = 150
            else:
                # Filter fits in one sheet
                ws = wb.create_sheet(title=filter_name[:31])

                color_hex = self.tab_colors[color_idx % len(self.tab_colors)][0]
                header_fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

                ws['A1'] = "Original Line #"
                ws['B1'] = "Log Content"
                ws['A1'].font = header_font
                ws['A1'].fill = header_fill
                ws['B1'].font = header_font
                ws['B1'].fill = header_fill

                for idx, (line_num, line) in enumerate(matches, start=2):
                    cell = ws[f'A{idx}']
                    cell.value = line_num
                    # Add hyperlink to All Logs
                    target_sheet, target_row = get_all_logs_location(line_num)
                    cell.hyperlink = f"#'{target_sheet}'!A{target_row}"
                    cell.font = link_font
                    ws[f'B{idx}'] = line

                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 150

            color_idx += 1

        # --- Summary Sheet ---
        ws_summary = wb.create_sheet(title="Summary")
        summary_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

        ws_summary['A1'] = "Metric"
        ws_summary['B1'] = "Value"
        ws_summary['A1'].font = header_font
        ws_summary['A1'].fill = summary_fill
        ws_summary['B1'].font = header_font
        ws_summary['B1'].fill = summary_fill

        row = 2
        ws_summary[f'A{row}'] = "Source File"
        ws_summary[f'B{row}'] = source_file
        row += 1
        ws_summary[f'A{row}'] = "Total Lines"
        ws_summary[f'B{row}'] = len(all_lines)
        row += 1
        ws_summary[f'A{row}'] = "All Logs Split Into"
        ws_summary[f'B{row}'] = f"{num_sheets} sheets"
        row += 2

        ws_summary[f'A{row}'] = "Filter Results"
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1

        for filter_name, data in filter_results.items():
            pct = len(data['matches']) / len(all_lines) * 100 if all_lines else 0
            ws_summary[f'A{row}'] = f"  {filter_name}"
            ws_summary[f'B{row}'] = f"{len(data['matches']):,} lines ({pct:.2f}%)"
            row += 1
            ws_summary[f'A{row}'] = f"    Keywords"
            ws_summary[f'B{row}'] = ", ".join(data['keywords'])
            row += 1

        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 80

        self.update_status("Saving Excel file...", "blue")
        wb.save(output_path)

    def create_excel_report_concat_split(self, sorted_tagged_lines, filter_results, output_path, source_files):
        """
        Create Excel report for concatenated multi-file analysis with splitting.
        Used when total lines exceed Excel's row limit.
        """
        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF")

        # Calculate sheets needed
        max_data_rows = EXCEL_MAX_ROWS - 1
        total_lines = len(sorted_tagged_lines)
        num_sheets = (total_lines + max_data_rows - 1) // max_data_rows

        self.update_status(f"Splitting All Logs across {num_sheets} sheets...", "blue")

        all_logs_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")

        # --- All Logs Sheets (split) ---
        # Build mapping from tagged line to (sheet_name, row) for hyperlinks
        all_logs_location_map = {}  # {id(tagged_line): (sheet_name, row)}

        for sheet_num in range(num_sheets):
            start_idx = sheet_num * max_data_rows
            end_idx = min((sheet_num + 1) * max_data_rows, total_lines)

            if sheet_num == 0:
                ws = wb.active
                ws.title = "All Logs"
                sheet_name = "All Logs"
            else:
                sheet_name = f"All Logs ({sheet_num + 1})"
                ws = wb.create_sheet(title=sheet_name)

            # Headers
            ws['A1'] = "Line #"
            ws['B1'] = "Date"
            ws['C1'] = "Source File"
            ws['D1'] = "Log Content"
            for col in ['A1', 'B1', 'C1', 'D1']:
                ws[col].font = header_font
                ws[col].fill = all_logs_fill

            # Add data rows and track locations
            for row_offset, line_idx in enumerate(range(start_idx, end_idx)):
                excel_row = row_offset + 2
                tagged_line = sorted_tagged_lines[line_idx]
                timestamp, filename, line, original_line_num = tagged_line
                line_date = str(timestamp.date()) if timestamp else "N/A"

                ws[f'A{excel_row}'] = line_idx + 1
                ws[f'B{excel_row}'] = line_date
                ws[f'C{excel_row}'] = filename
                ws[f'D{excel_row}'] = line

                # Track location for hyperlinks
                all_logs_location_map[id(tagged_line)] = (sheet_name, excel_row)

            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 120

            self.update_status(f"Wrote All Logs sheet {sheet_num + 1}/{num_sheets}...", "blue")

        # --- Filter Sheets ---
        color_idx = 0
        link_font = Font(color="0563C1", underline="single")  # Blue underlined hyperlink style

        for filter_name, data in filter_results.items():
            matches = data['matches']
            filter_total = len(matches)

            color_hex = self.tab_colors[color_idx % len(self.tab_colors)][0]
            header_fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

            if filter_total > max_data_rows:
                filter_num_sheets = (filter_total + max_data_rows - 1) // max_data_rows
                for sheet_num in range(filter_num_sheets):
                    start_idx = sheet_num * max_data_rows
                    end_idx = min((sheet_num + 1) * max_data_rows, filter_total)

                    if sheet_num == 0:
                        sheet_title = filter_name[:31]
                    else:
                        base_title = filter_name[:26]
                        sheet_title = f"{base_title} ({sheet_num + 1})"

                    ws = wb.create_sheet(title=sheet_title)

                    ws['A1'] = "Line #"
                    ws['B1'] = "Date"
                    ws['C1'] = "Source File"
                    ws['D1'] = "Log Content"
                    for col in ['A1', 'B1', 'C1', 'D1']:
                        ws[col].font = header_font
                        ws[col].fill = header_fill

                    for row_offset, match_idx in enumerate(range(start_idx, end_idx)):
                        excel_row = row_offset + 2
                        tagged_line = matches[match_idx]
                        timestamp, filename, line, original_line_num = tagged_line
                        line_date = str(timestamp.date()) if timestamp else "N/A"

                        cell = ws[f'A{excel_row}']
                        cell.value = row_offset + 1
                        # Add hyperlink to All Logs
                        location = all_logs_location_map.get(id(tagged_line))
                        if location:
                            target_sheet, target_row = location
                            cell.hyperlink = f"#'{target_sheet}'!A{target_row}"
                            cell.font = link_font
                        ws[f'B{excel_row}'] = line_date
                        ws[f'C{excel_row}'] = filename
                        ws[f'D{excel_row}'] = line

                    ws.column_dimensions['A'].width = 10
                    ws.column_dimensions['B'].width = 12
                    ws.column_dimensions['C'].width = 30
                    ws.column_dimensions['D'].width = 120
            else:
                ws = wb.create_sheet(title=filter_name[:31])

                ws['A1'] = "Line #"
                ws['B1'] = "Date"
                ws['C1'] = "Source File"
                ws['D1'] = "Log Content"
                for col in ['A1', 'B1', 'C1', 'D1']:
                    ws[col].font = header_font
                    ws[col].fill = header_fill

                for idx, tagged_line in enumerate(matches, start=1):
                    excel_row = idx + 1
                    timestamp, filename, line, original_line_num = tagged_line
                    line_date = str(timestamp.date()) if timestamp else "N/A"

                    cell = ws[f'A{excel_row}']
                    cell.value = idx
                    # Add hyperlink to All Logs
                    location = all_logs_location_map.get(id(tagged_line))
                    if location:
                        target_sheet, target_row = location
                        cell.hyperlink = f"#'{target_sheet}'!A{target_row}"
                        cell.font = link_font
                    ws[f'B{excel_row}'] = line_date
                    ws[f'C{excel_row}'] = filename
                    ws[f'D{excel_row}'] = line

                ws.column_dimensions['A'].width = 10
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 30
                ws.column_dimensions['D'].width = 120

            color_idx += 1

        # --- Summary Sheet ---
        ws_summary = wb.create_sheet(title="Summary")
        summary_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

        ws_summary['A1'] = "Metric"
        ws_summary['B1'] = "Value"
        ws_summary['A1'].font = header_font
        ws_summary['A1'].fill = summary_fill
        ws_summary['B1'].font = header_font
        ws_summary['B1'].fill = summary_fill

        unique_dates = self.extract_unique_dates(sorted_tagged_lines)

        row = 2
        ws_summary[f'A{row}'] = "Mode"
        ws_summary[f'B{row}'] = "Multi-File Concatenation"
        row += 1
        ws_summary[f'A{row}'] = "Number of Files"
        ws_summary[f'B{row}'] = len(source_files)
        row += 1
        ws_summary[f'A{row}'] = "Total Lines"
        ws_summary[f'B{row}'] = total_lines
        row += 1
        ws_summary[f'A{row}'] = "All Logs Split Into"
        ws_summary[f'B{row}'] = f"{num_sheets} sheets"
        row += 2

        ws_summary[f'A{row}'] = "Source Files"
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1
        for file_path in source_files:
            ws_summary[f'A{row}'] = f"  {os.path.basename(file_path)}"
            row += 1

        if unique_dates:
            row += 1
            ws_summary[f'A{row}'] = "Date Range"
            ws_summary[f'A{row}'].font = Font(bold=True)
            row += 1
            ws_summary[f'A{row}'] = "  Start Date"
            ws_summary[f'B{row}'] = str(unique_dates[0])
            row += 1
            ws_summary[f'A{row}'] = "  End Date"
            ws_summary[f'B{row}'] = str(unique_dates[-1])
            row += 1

        row += 1
        ws_summary[f'A{row}'] = "Filter Results"
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1

        for filter_name, data in filter_results.items():
            ws_summary[f'A{row}'] = f"  {filter_name}"
            ws_summary[f'B{row}'] = f"{len(data['matches']):,} lines"
            row += 1
            ws_summary[f'A{row}'] = f"    Keywords"
            ws_summary[f'B{row}'] = ", ".join(data['keywords'])
            row += 1

        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 80

        self.update_status("Saving Excel file...", "blue")
        wb.save(output_path)


class FilterNameDialog:
    """Simple dialog to get a filter name."""

    def __init__(self, parent, title, existing_names=None):
        self.result = None
        self.existing_names = existing_names or []

        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("350x120")
        self.dialog.resizable(False, False)
        self.dialog.transient(parent)
        self.dialog.grab_set()

        # Center on parent
        self.dialog.geometry(f"+{parent.winfo_x() + 150}+{parent.winfo_y() + 200}")

        frame = ttk.Frame(self.dialog, padding="20")
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="Enter tab name:").pack(anchor="w")

        self.entry = ttk.Entry(frame, width=40)
        self.entry.pack(fill="x", pady=(5, 15))
        self.entry.focus_set()
        self.entry.bind("<Return>", lambda e: self.ok())

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill="x")

        ttk.Button(btn_frame, text="OK", command=self.ok, width=10).pack(side="right", padx=(5, 0))
        ttk.Button(btn_frame, text="Cancel", command=self.cancel, width=10).pack(side="right")

        self.dialog.wait_window()

    def ok(self):
        name = self.entry.get().strip()
        if not name:
            messagebox.showwarning("Empty", "Please enter a tab name.", parent=self.dialog)
            return
        if name in self.existing_names:
            messagebox.showwarning("Duplicate", f"'{name}' already exists.", parent=self.dialog)
            return
        # Check for invalid Excel sheet name characters
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        if any(c in name for c in invalid_chars):
            messagebox.showwarning("Invalid Name", f"Tab name cannot contain: {' '.join(invalid_chars)}", parent=self.dialog)
            return
        self.result = name
        self.dialog.destroy()

    def cancel(self):
        self.dialog.destroy()


class RowLimitDialog:
    """Dialog to handle Excel row limit exceeded scenario."""

    def __init__(self, parent, total_rows):
        self.result = None  # Will be "csv", "split", or None (cancelled)

        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Row Limit Exceeded")
        self.dialog.geometry("450x200")
        self.dialog.resizable(False, False)
        self.dialog.transient(parent)
        self.dialog.grab_set()

        # Center on parent
        self.dialog.geometry(f"+{parent.winfo_x() + 100}+{parent.winfo_y() + 150}")

        frame = ttk.Frame(self.dialog, padding="20")
        frame.pack(fill="both", expand=True)

        # Warning message
        excess = total_rows - EXCEL_MAX_ROWS + 1  # +1 for header row
        warning_text = (
            f"Your log file contains {total_rows:,} lines, which exceeds\n"
            f"Excel's maximum of {EXCEL_MAX_ROWS:,} rows per sheet.\n\n"
            f"How would you like to proceed?"
        )
        ttk.Label(frame, text=warning_text, justify="center").pack(pady=(0, 15))

        # Option buttons
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill="x", pady=(10, 0))

        csv_btn = ttk.Button(
            btn_frame,
            text="Export as CSV",
            command=self.choose_csv,
            width=20
        )
        csv_btn.pack(side="left", padx=(20, 10))

        split_btn = ttk.Button(
            btn_frame,
            text="Split Across Tabs",
            command=self.choose_split,
            width=20
        )
        split_btn.pack(side="left", padx=(10, 10))

        cancel_btn = ttk.Button(
            btn_frame,
            text="Cancel",
            command=self.cancel,
            width=10
        )
        cancel_btn.pack(side="right", padx=(10, 20))

        self.dialog.wait_window()

    def choose_csv(self):
        self.result = "csv"
        self.dialog.destroy()

    def choose_split(self):
        self.result = "split"
        self.dialog.destroy()

    def cancel(self):
        self.result = None
        self.dialog.destroy()


def main():
    # Use TkinterDnD for drag and drop support if available
    if DND_AVAILABLE:
        root = TkinterDnD.Tk()
    else:
        root = tk.Tk()
    app = LogAnalyzerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
