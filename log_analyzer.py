"""
Log Analyzer - Reads log files and exports to Excel with filtered error lines.

Usage:
    python log_analyzer.py <log_file_path> [output_file_path]

Example:
    python log_analyzer.py "C:\logs\mylog.log"
    python log_analyzer.py "C:\logs\mylog.log" "C:\output\results.xlsx"
"""

import sys
import os
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Error: openpyxl is not installed.")
    print("Please install it by running: pip install openpyxl")
    sys.exit(1)


# =============================================================================
# CONFIGURABLE ERROR KEYWORDS
# Add or remove keywords as needed. Lines containing ANY of these (case-insensitive)
# will be flagged as error lines.
# =============================================================================
ERROR_KEYWORDS = [
    "Error",
    "Exception",
    "Fatal",
    "Critical",
    "Fail",
    "Failed",
    # Add more keywords below as needed:
    # "Warning",
    # "Timeout",
]


def read_log_file(file_path: str) -> list[str]:
    """Read all lines from a log file."""
    encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']

    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                lines = f.readlines()
            print(f"Successfully read file using {encoding} encoding.")
            return [line.rstrip('\n\r') for line in lines]
        except UnicodeDecodeError:
            continue

    raise ValueError(f"Could not read file with any of the attempted encodings: {encodings}")


def filter_error_lines(lines: list[str], keywords: list[str]) -> list[tuple[int, str]]:
    """
    Filter lines that contain any of the error keywords.
    Returns a list of tuples: (line_number, line_content)
    """
    error_lines = []
    keywords_lower = [kw.lower() for kw in keywords]

    for line_num, line in enumerate(lines, start=1):
        line_lower = line.lower()
        if any(keyword in line_lower for keyword in keywords_lower):
            error_lines.append((line_num, line))

    return error_lines


def create_excel_report(
    all_lines: list[str],
    error_lines: list[tuple[int, str]],
    output_path: str,
    source_file: str
):
    """Create an Excel workbook with two sheets: All Logs and Errors."""
    wb = Workbook()

    # --- Sheet 1: All Logs ---
    ws_all = wb.active
    ws_all.title = "All Logs"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Add headers
    ws_all['A1'] = "Line #"
    ws_all['B1'] = "Log Content"
    ws_all['A1'].font = header_font
    ws_all['A1'].fill = header_fill
    ws_all['B1'].font = header_font
    ws_all['B1'].fill = header_fill

    # Add all log lines
    for idx, line in enumerate(all_lines, start=2):
        ws_all[f'A{idx}'] = idx - 1  # Line number
        ws_all[f'B{idx}'] = line

    # Adjust column widths
    ws_all.column_dimensions['A'].width = 10
    ws_all.column_dimensions['B'].width = 150

    # --- Sheet 2: Errors ---
    ws_errors = wb.create_sheet(title="Errors")

    # Error header styling (red theme)
    error_header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    # Add headers
    ws_errors['A1'] = "Original Line #"
    ws_errors['B1'] = "Log Content"
    ws_errors['A1'].font = header_font
    ws_errors['A1'].fill = error_header_fill
    ws_errors['B1'].font = header_font
    ws_errors['B1'].fill = error_header_fill

    # Add error lines
    for idx, (line_num, line) in enumerate(error_lines, start=2):
        ws_errors[f'A{idx}'] = line_num
        ws_errors[f'B{idx}'] = line

    # Adjust column widths
    ws_errors.column_dimensions['A'].width = 15
    ws_errors.column_dimensions['B'].width = 150

    # --- Sheet 3: Summary ---
    ws_summary = wb.create_sheet(title="Summary")

    summary_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    ws_summary['A1'] = "Metric"
    ws_summary['B1'] = "Value"
    ws_summary['A1'].font = header_font
    ws_summary['A1'].fill = summary_header_fill
    ws_summary['B1'].font = header_font
    ws_summary['B1'].fill = summary_header_fill

    ws_summary['A2'] = "Source File"
    ws_summary['B2'] = source_file
    ws_summary['A3'] = "Total Lines"
    ws_summary['B3'] = len(all_lines)
    ws_summary['A4'] = "Error Lines Found"
    ws_summary['B4'] = len(error_lines)
    ws_summary['A5'] = "Error Percentage"
    ws_summary['B5'] = f"{(len(error_lines) / len(all_lines) * 100):.2f}%" if all_lines else "0%"
    ws_summary['A6'] = "Keywords Used"
    ws_summary['B6'] = ", ".join(ERROR_KEYWORDS)

    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 80

    # Save the workbook
    wb.save(output_path)
    print(f"\nExcel report saved to: {output_path}")


def main():
    # Parse command line arguments
    if len(sys.argv) < 2:
        print("Usage: python log_analyzer.py <log_file_path> [output_file_path]")
        print("\nExample:")
        print('  python log_analyzer.py "C:\\logs\\mylog.log"')
        print('  python log_analyzer.py "C:\\logs\\mylog.log" "C:\\output\\results.xlsx"')
        sys.exit(1)

    input_file = sys.argv[1]

    # Validate input file exists
    if not os.path.isfile(input_file):
        print(f"Error: File not found: {input_file}")
        sys.exit(1)

    # Determine output file path
    if len(sys.argv) >= 3:
        output_file = sys.argv[2]
    else:
        # Default: same directory as input, with .xlsx extension
        input_path = Path(input_file)
        output_file = str(input_path.parent / f"{input_path.stem}_analysis.xlsx")

    print(f"Log Analyzer")
    print(f"{'=' * 50}")
    print(f"Input file:  {input_file}")
    print(f"Output file: {output_file}")
    print(f"Keywords:    {', '.join(ERROR_KEYWORDS)}")
    print(f"{'=' * 50}")

    # Read the log file
    print("\nReading log file...")
    all_lines = read_log_file(input_file)
    print(f"Total lines read: {len(all_lines):,}")

    # Filter error lines
    print("\nFiltering error lines...")
    error_lines = filter_error_lines(all_lines, ERROR_KEYWORDS)
    print(f"Error lines found: {len(error_lines):,}")

    # Create Excel report
    print("\nGenerating Excel report...")
    create_excel_report(all_lines, error_lines, output_file, input_file)

    print(f"\nDone! Found {len(error_lines):,} error lines out of {len(all_lines):,} total lines.")


if __name__ == "__main__":
    main()
